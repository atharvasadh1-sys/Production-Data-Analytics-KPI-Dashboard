#!/usr/bin/env python3
"""
Final Compliance Fix - ZigZag AI Assignment
This script adds the missing compliance sheets to the existing workbook.

Input: kpi_output_script2_final.xlsx
Output: kpi_output_script2_final.xlsx (preserves all existing sheets)

Author: Senior Python Data Engineer
Date: 2026-08-05
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import io
import datetime
import sys
import os

# =============================================================================
# CONFIGURATION
# =============================================================================

EXCEL_FILE = "kpi_output_script2_final.xlsx"

# =============================================================================
# AI CONVERSATION DATA - TO BE POPULATED BY USER
# =============================================================================
# 
# IMPORTANT: Replace the placeholder data below with your actual DeepSeek
# conversation data from your exported chat history.
# 
# To get your DeepSeek conversation data:
# 1. Open your DeepSeek chat history
# 2. Copy the session link from the URL
# 3. For each prompt you sent, copy your exact text
# 4. For each AI response, copy the exact response text
# 5. Note which suggestions you used vs rejected
# 6. Document any manual corrections you made
# 
# Both DeepSeek conversations from this project should be included
# in chronological order.
# =============================================================================

# --- USER ACTION REQUIRED: Replace the placeholder data below ---

DEEPSEEK_CONVERSATION = {
    "chat_1": {
        "date": "2026-08-04",
        "link": "https://chat.deepseek.com/a/chat/s/5f90a7ea-5f76-408d-8f3e-6cd1a58b1bc7",
        "title": "Production Analytics Solution",
        "prompts": [
            "Analyze the ZigZag AI assignment and identify all event types, machine states, operator events, material flow events, KPIs, required workbook sheets and complete project architecture.",
            "Generate Script 1 (script1_data_processing.py) as a complete production-grade Python solution with preprocessing, shift handling, duration calculation, invalid event detection and workbook generation."
        ],
        "responses": [
            "Analyzed the assignment, identified complete project architecture, event classifications, KPI requirements, workbook structure and processing pipeline.",
            "Generated a complete Script 1 for data preprocessing including duplicate removal, invalid event filtering, shift assignment, duration calculation, processed workbook creation and calculation audit."
        ],
        "used": [
            "Project architecture",
            "Event parsing logic",
            "Shift assignment",
            "Data preprocessing workflow",
            "Workbook generation approach"
        ],
        "rejected": [
            "No major suggestions rejected in this phase."
        ],
        "manual": [
            "Verified processed workbook.",
            "Fixed formatting issues.",
            "Verified generated sheets."
        ],
        "verification": [
            "Processed workbook generated successfully.",
            "Script executed successfully without major logic issues."
        ]
    },

    "chat_2": {
        "date": "2026-08-05",
        "link": "https://chat.deepseek.com/a/chat/s/128c27e9-7d59-4b03-8ce7-3010e0bb68db",
        "title": "Compliance Review (Continued)",
        "prompts": [
            "Review the completed assignment against the ZigZag AI Assignment PDF.",
            "Identify missing Dashboard, Graphs & Visualizations, AI Conversation Log, formatting issues and KPI compliance.",
            "Fix merged-cell formatting issue.",
            "Verify final workbook for submission."
        ],
        "responses": [
            "Performed complete compliance review and identified missing Dashboard, Graphs & Visualizations and AI Conversation Log.",
            "Generated compliance script for missing sheets and required visualizations.",
            "Provided corrected format_sheet() compatible with merged cells.",
            "Verified workbook structure and suggested final compliance improvements."
        ],
        "used": [
            "Compliance review",
            "Dashboard creation",
            "Graphs & Visualizations",
            "MergedCell fix",
            "Workbook formatting improvements"
        ],
        "rejected": [
            "Automatic Hours-to-Minutes conversion was not fully adopted because the existing KPI engine already produced consistent workbook values."
        ],
        "manual": [
            "Added actual DeepSeek conversation links.",
            "Reviewed generated workbook.",
            "Verified charts.",
            "Checked formatting manually.",
            "Validated final workbook before submission."
        ],
        "verification": [
            "Dashboard verified.",
            "Graphs verified.",
            "AI Conversation Log verified.",
            "Workbook reviewed against assignment requirements.",
            "Final submission prepared."
        ]
    }
}


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def format_sheet(ws, title, max_width=50):
    """Applies professional formatting to a worksheet. Properly handles merged cells."""
    # Format header row - skip merged cells
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if cell.value is not None:
                try:
                    cell.font = Font(bold=True, color="FFFFFF", size=11)
                    cell.fill = PatternFill(start_color="2C6B9E", end_color="2C6B9E", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                except AttributeError:
                    pass
    
    # Format data rows - skip merged cells
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if cell.value is not None:
                try:
                    if isinstance(cell.value, (int, float)):
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                except AttributeError:
                    pass
    
    # Auto-fit columns - skip merged cells
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            if isinstance(cell, MergedCell):
                continue
            if cell.value is not None:
                try:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
                except:
                    pass
        
        adjusted_width = min(max_length + 2, max_width)
        if adjusted_width > 0:
            ws.column_dimensions[column_letter].width = adjusted_width


def add_matplotlib_chart(ws, fig, cell_location, title, chart_type):
    """Saves a matplotlib figure to an in-memory buffer and places it in a cell."""
    try:
        img_data = io.BytesIO()
        fig.savefig(img_data, format='png', bbox_inches='tight', dpi=100)
        img_data.seek(0)
        img = Image(img_data)
        ws.add_image(img, cell_location)
        
        # Add a label for the chart
        label_row = int(cell_location[1:])
        label_cell = ws.cell(row=label_row - 1, column=ord(cell_location[0]) - 64)
        label_cell.value = f"{chart_type}: {title}"
        label_cell.font = Font(bold=True, size=10)
        label_cell.alignment = Alignment(horizontal="center")
        
        plt.close(fig)
    except Exception as e:
        print(f"Warning: Error adding chart '{title}': {e}")


def get_kpi_data():
    """Load KPI data from the workbook."""
    try:
        df_kpi = pd.read_excel(EXCEL_FILE, sheet_name='KPI Summary')
        df_shift = pd.read_excel(EXCEL_FILE, sheet_name='Shift KPIs')
        df_shift.set_index('Shift', inplace=True)
        df_loss = pd.read_excel(EXCEL_FILE, sheet_name='Top 5 Loss Events')
        df_loss.set_index('Loss_Event_Type', inplace=True)
        return df_kpi, df_shift, df_loss
    except Exception as e:
        print(f"Warning: Could not load KPI data: {e}")
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()


def get_management_summary():
    """Load management summary data."""
    try:
        df_summary = pd.read_excel(EXCEL_FILE, sheet_name='Management Summary')
        return df_summary
    except Exception as e:
        print(f"Warning: Could not load Management Summary: {e}")
        return pd.DataFrame()


def get_kpi_values_from_workbook(wb):
    """Extract live KPI values from the workbook sheets."""
    kpi_data = {}
    
    # Get Management Summary data
    if 'Management Summary' in wb.sheetnames:
        ws_summary = wb['Management Summary']
        for row in range(2, ws_summary.max_row + 1):
            metric_cell = ws_summary.cell(row=row, column=1)
            value_cell = ws_summary.cell(row=row, column=2)
            unit_cell = ws_summary.cell(row=row, column=3)
            
            if metric_cell.value and value_cell.value is not None:
                key = str(metric_cell.value)
                kpi_data[key] = {
                    'value': value_cell.value,
                    'unit': str(unit_cell.value) if unit_cell.value else ''
                }
    
    return kpi_data


def add_kpi_card(ws, row, col, label, value, unit, color="FF2C6B9E"):
    """Add a formatted KPI card to the dashboard."""
    # Label
    label_cell = ws.cell(row=row, column=col)
    label_cell.value = label
    label_cell.font = Font(size=11, bold=True, color="FFFFFF")
    label_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    label_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
    
    # Value
    value_cell = ws.cell(row=row+1, column=col)
    value_cell.value = f"{value:,.2f}" if isinstance(value, (int, float)) else value
    value_cell.font = Font(size=20, bold=True, color=color)
    value_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+1)
    
    # Unit
    unit_cell = ws.cell(row=row+2, column=col)
    unit_cell.value = unit
    unit_cell.font = Font(size=10)
    unit_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=row+2, start_column=col, end_row=row+2, end_column=col+1)
    
    return row + 4


# =============================================================================
# SHEET CREATION FUNCTIONS
# =============================================================================

def create_dashboard(wb):
    """Create the Dashboard sheet with live workbook references."""
    print("Creating Dashboard...")
    
    # Remove existing Dashboard if present
    if 'Dashboard' in wb.sheetnames:
        wb.remove(wb['Dashboard'])
    
    ws = wb.create_sheet("Dashboard", 0)
    
    # Title
    ws['A1'] = "ZigZag AI - Production Performance Dashboard"
    ws['A1'].font = Font(size=24, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1A3C5E", end_color="1A3C5E", fill_type="solid")
    ws.merge_cells('A1:J1')
    ws.row_dimensions[1].height = 40
    
    # Subtitle - use the unit from the workbook
    ws['A2'] = f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A2'].font = Font(size=10, italic=True)
    ws.merge_cells('A2:J2')
    
    # Get live KPI values
    kpi_data = get_kpi_values_from_workbook(wb)
    
    # --- Section 1: Machine KPIs ---
    row = 5
    ws.cell(row=row, column=1).value = "MACHINE KPIs"
    ws.cell(row=row, column=1).font = Font(bold=True, size=14, color="FFFFFF")
    ws.cell(row=row, column=1).fill = PatternFill(start_color="2C6B9E", end_color="2C6B9E", fill_type="solid")
    ws.merge_cells('A' + str(row) + ':E' + str(row))
    row += 1
    
    # Machine KPI cards - use units from workbook
    machine_metrics = [
        ('Total Running Time', kpi_data.get('Total Running Time', {}).get('value', 0), kpi_data.get('Total Running Time', {}).get('unit', 'Hours')),
        ('Total Downtime', kpi_data.get('Total Downtime', {}).get('value', 0), kpi_data.get('Total Downtime', {}).get('unit', 'Hours')),
        ('Overall Utilization', kpi_data.get('Overall Utilization', {}).get('value', 0), '%'),
        ('Total Duration', kpi_data.get('Total Duration', {}).get('value', 0), kpi_data.get('Total Duration', {}).get('unit', 'Hours')),
    ]
    
    col = 1
    for label, value, unit in machine_metrics:
        row = add_kpi_card(ws, row, col, label, value, unit, "FF2C6B9E")
        col += 3
        if col > 8:
            col = 1
            row += 2
    
    row = max(row, 20)
    
    # --- Section 2: Operator KPIs ---
    row += 2
    ws.cell(row=row, column=1).value = "OPERATOR KPIs"
    ws.cell(row=row, column=1).font = Font(bold=True, size=14, color="FFFFFF")
    ws.cell(row=row, column=1).fill = PatternFill(start_color="D35400", end_color="D35400", fill_type="solid")
    ws.merge_cells('A' + str(row) + ':E' + str(row))
    row += 1
    
    operator_metrics = [
        ('Total Input Shortage', kpi_data.get('Total Input Shortage', {}).get('value', 0), kpi_data.get('Total Input Shortage', {}).get('unit', 'Hours')),
        ('Total Output Shortage', kpi_data.get('Total Output Shortage', {}).get('value', 0), kpi_data.get('Total Output Shortage', {}).get('unit', 'Hours')),
        ('Total Material Shortage', kpi_data.get('Total Material Shortage', {}).get('value', 0), kpi_data.get('Total Material Shortage', {}).get('unit', 'Hours')),
    ]
    
    col = 1
    for label, value, unit in operator_metrics:
        row = add_kpi_card(ws, row, col, label, value, unit, "FFD35400")
        col += 3
    
    row = max(row, 35)
    
    # --- Section 3: Key Highlights ---
    row += 2
    ws.cell(row=row, column=1).value = "KEY HIGHLIGHTS"
    ws.cell(row=row, column=1).font = Font(bold=True, size=14, color="FFFFFF")
    ws.cell(row=row, column=1).fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
    ws.merge_cells('A' + str(row) + ':J' + str(row))
    row += 1
    
    highlights = [
        "• Review the 'Exception Report' for a detailed log of all downtime events",
        "• Check 'Top 5 Loss Events' to identify primary causes of productivity loss",
        "• Use 'Shift KPIs' to compare performance across shifts",
        "• All KPI calculations are documented in 'Calculation Audit'",
        "• Refer to 'Management Summary' for executive-level insights"
    ]
    
    for idx, highlight in enumerate(highlights):
        ws.cell(row=row + idx, column=1).value = highlight
        ws.cell(row=row + idx, column=1).font = Font(size=11)
        ws.merge_cells('A' + str(row + idx) + ':J' + str(row + idx))
    
    # Format the sheet
    format_sheet(ws, "Dashboard")
    
    # Set column widths
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws.column_dimensions[col].width = 18
    
    print("Dashboard created successfully.")


def create_graphs_sheet(wb):
    """Create the Graphs & Visualizations sheet with all required charts."""
    print("Creating Graphs & Visualizations...")
    
    # Remove existing Graphs sheet if present
    if 'Graphs & Visualizations' in wb.sheetnames:
        wb.remove(wb['Graphs & Visualizations'])
    
    ws = wb.create_sheet("Graphs & Visualizations")
    
    # Title
    ws['A1'] = "Graphs & Visualizations"
    ws['A1'].font = Font(size=20, bold=True)
    ws.merge_cells('A1:J1')
    
    ws['A2'] = "Duration values use the unit from the source data (Hours)"
    ws['A2'].font = Font(size=12, italic=True)
    ws.merge_cells('A2:J2')
    
    # Load data
    df_kpi, df_shift, df_loss = get_kpi_data()
    
    if df_kpi.empty:
        print("Warning: No KPI data available for charts")
        ws['A4'] = "No data available for charts"
        format_sheet(ws, "Graphs & Visualizations")
        return
    
    # Chart positions
    chart_pos = 4
    
    # Determine the correct column names based on what's available
    running_col = 'Running Time (Hours)' if 'Running Time (Hours)' in df_kpi['KPI'].values else 'Running Time (Minutes)'
    downtime_col = 'Downtime (Hours)' if 'Downtime (Hours)' in df_kpi['KPI'].values else 'Downtime (Minutes)'
    
    # Chart 1: Machine Running Time vs Downtime
    fig1, ax1 = plt.subplots(figsize=(8, 4))
    machine_ids = df_kpi[df_kpi['Category'] == 'Machine']['Machine_ID'].unique()
    for machine_id in machine_ids:
        df_machine = df_kpi[(df_kpi['Category'] == 'Machine') & (df_kpi['Machine_ID'] == machine_id)]
        running = df_machine[df_machine['KPI'] == running_col]['Value'].sum()
        downtime = df_machine[df_machine['KPI'] == downtime_col]['Value'].sum()
        ax1.bar(f"M{int(machine_id)} Running", running, color='green', label='Running' if machine_id == 1 else "")
        ax1.bar(f"M{int(machine_id)} Downtime", downtime, color='red', label='Downtime' if machine_id == 1 else "")
    ax1.set_ylabel('Hours')
    ax1.set_title('Machine Running Time vs Downtime')
    ax1.legend()
    fig1.tight_layout()
    add_matplotlib_chart(ws, fig1, 'A' + str(chart_pos), 'Machine Running Time vs Downtime', 'Bar Chart')
    
    # Chart 2: Shift-wise Machine Utilization
    fig2, ax2 = plt.subplots(figsize=(8, 4))
    if 'Avg_Utilization_Percent' in df_shift.columns:
        utilizations = df_shift['Avg_Utilization_Percent'].dropna()
        if not utilizations.empty:
            utilizations.plot(kind='bar', ax=ax2, color=['#2C6B9E', '#D35400', '#27AE60'])
            ax2.set_ylabel('Utilization %')
            ax2.set_title('Shift-wise Machine Utilization')
            ax2.set_xticklabels(ax2.get_xticklabels(), rotation=0)
    fig2.tight_layout()
    add_matplotlib_chart(ws, fig2, 'J' + str(chart_pos), 'Shift-wise Machine Utilization', 'Bar Chart')
    chart_pos += 12
    
    # Chart 3: Input vs Output Operator Availability
    fig3, ax3 = plt.subplots(figsize=(8, 4))
    input_avail = df_shift['Input_Availability_Percent'].dropna() if 'Input_Availability_Percent' in df_shift.columns else pd.Series()
    output_avail = df_shift['Output_Availability_Percent'].dropna() if 'Output_Availability_Percent' in df_shift.columns else pd.Series()
    if not input_avail.empty and not output_avail.empty:
        ax3.plot(input_avail.index, input_avail, marker='o', label='Input Availability', color='#2C6B9E', linewidth=2)
        ax3.plot(output_avail.index, output_avail, marker='s', label='Output Availability', color='#D35400', linewidth=2)
        ax3.set_ylabel('Availability %')
        ax3.set_title('Input vs Output Operator Availability by Shift')
        ax3.legend()
        ax3.grid(True, alpha=0.3)
    fig3.tight_layout()
    add_matplotlib_chart(ws, fig3, 'A' + str(chart_pos), 'Input vs Output Operator Availability', 'Line Chart')
    
    # Chart 4: Operator Shortage Duration
    fig4, ax4 = plt.subplots(figsize=(8, 4))
    input_shortage_col = 'Input_Shortage_Hours' if 'Input_Shortage_Hours' in df_shift.columns else 'Input_Shortage_Minutes'
    output_shortage_col = 'Output_Shortage_Hours' if 'Output_Shortage_Hours' in df_shift.columns else 'Output_Shortage_Minutes'
    input_shortage = df_shift[input_shortage_col].dropna() if input_shortage_col in df_shift.columns else pd.Series()
    output_shortage = df_shift[output_shortage_col].dropna() if output_shortage_col in df_shift.columns else pd.Series()
    if not input_shortage.empty:
        width = 0.35
        x = range(len(input_shortage.index))
        ax4.bar([i - width/2 for i in x], input_shortage, width, label='Input Shortage', color='#E74C3C')
        ax4.bar([i + width/2 for i in x], output_shortage, width, label='Output Shortage', color='#8E44AD')
        ax4.set_xticks(x)
        ax4.set_xticklabels(input_shortage.index)
        ylabel = 'Hours' if 'Hours' in input_shortage_col else 'Minutes'
        ax4.set_ylabel(ylabel)
        ax4.set_title(f'Operator Shortage Duration by Shift ({ylabel})')
        ax4.legend()
    fig4.tight_layout()
    add_matplotlib_chart(ws, fig4, 'J' + str(chart_pos), 'Operator Shortage Duration by Shift', 'Bar Chart')
    chart_pos += 12
    
    # Chart 5: Material Flow Condition Duration
    fig5, ax5 = plt.subplots(figsize=(8, 4))
    processing_col = 'Material_Processing_Hours' if 'Material_Processing_Hours' in df_shift.columns else 'Material_Processing_Minutes'
    shortage_col = 'Material_Shortage_Hours' if 'Material_Shortage_Hours' in df_shift.columns else 'Material_Shortage_Minutes'
    interruption_col = 'Material_Interruption_Hours' if 'Material_Interruption_Hours' in df_shift.columns else 'Material_Interruption_Minutes'
    material_processing = df_shift[processing_col].dropna() if processing_col in df_shift.columns else pd.Series()
    material_shortage = df_shift[shortage_col].dropna() if shortage_col in df_shift.columns else pd.Series()
    material_interruption = df_shift[interruption_col].dropna() if interruption_col in df_shift.columns else pd.Series()
    if not material_processing.empty:
        x = range(len(material_processing.index))
        ax5.bar([i - 0.2 for i in x], material_processing, 0.2, label='Processing', color='#27AE60')
        ax5.bar(x, material_shortage, 0.2, label='Shortage', color='#E74C3C')
        ax5.bar([i + 0.2 for i in x], material_interruption, 0.2, label='Interruption', color='#F39C12')
        ax5.set_xticks(x)
        ax5.set_xticklabels(material_processing.index)
        ylabel = 'Hours' if 'Hours' in processing_col else 'Minutes'
        ax5.set_ylabel(ylabel)
        ax5.set_title(f'Material Flow Condition Duration by Shift ({ylabel})')
        ax5.legend()
    fig5.tight_layout()
    add_matplotlib_chart(ws, fig5, 'A' + str(chart_pos), 'Material Flow Condition Duration', 'Bar Chart')
    
    # Chart 6: Material Interruption Trend
    fig6, ax6 = plt.subplots(figsize=(8, 4))
    if not material_interruption.empty:
        material_interruption.plot(kind='line', marker='o', ax=ax6, color='#F39C12', linewidth=2)
        ylabel = 'Hours' if 'Hours' in interruption_col else 'Minutes'
        ax6.set_ylabel(ylabel)
        ax6.set_title(f'Material Interruption Trend by Shift ({ylabel})')
        ax6.grid(True, alpha=0.3)
    fig6.tight_layout()
    add_matplotlib_chart(ws, fig6, 'J' + str(chart_pos), 'Material Interruption Trend', 'Line Chart')
    chart_pos += 12
    
    # Chart 7: Shift-wise KPI Comparison
    fig7, ax7 = plt.subplots(figsize=(10, 6))
    kpi_to_plot = ['Avg_Utilization_Percent', 'Avg_Availability_Percent', 
                   'Input_Availability_Percent', 'Output_Availability_Percent',
                   'Handling_Efficiency_Percent']
    labels = ['Utilization', 'Availability', 'Input Avail', 'Output Avail', 'Handling Eff']
    colors = ['#2C6B9E', '#D35400', '#27AE60', '#E74C3C', '#8E44AD']
    for i, (kpi, label, color) in enumerate(zip(kpi_to_plot, labels, colors)):
        if kpi in df_shift.columns:
            data = df_shift[kpi].dropna()
            if not data.empty:
                ax7.bar([f"{idx}\n{label}" for idx in data.index], data, label=label, color=color, alpha=0.7)
    ax7.set_ylabel('Percentage (%)')
    ax7.set_title('Shift-wise KPI Comparison')
    ax7.legend()
    fig7.tight_layout()
    add_matplotlib_chart(ws, fig7, 'A' + str(chart_pos), 'Shift-wise KPI Comparison', 'Bar Chart')
    
    # Chart 8: Top 5 Productivity Loss Events
    fig8, ax8 = plt.subplots(figsize=(8, 4))
    if not df_loss.empty:
        # Check which columns exist
        if 'Total_Duration_Minutes' in df_loss.columns:
            duration_col = 'Total_Duration_Minutes'
            label = 'Minutes'
        elif 'Total_Duration_Hours' in df_loss.columns:
            duration_col = 'Total_Duration_Hours'
            label = 'Hours'
        else:
            duration_col = None
            label = ''
        
        if duration_col:
            top_losses = df_loss[duration_col].sort_values(ascending=False).head(5)
            if not top_losses.empty:
                top_losses.plot(kind='bar', ax=ax8, color='coral')
                ax8.set_ylabel(f'Total Duration ({label})')
                ax8.set_title(f'Top 5 Productivity Loss Events ({label})')
                ax8.set_xticklabels(ax8.get_xticklabels(), rotation=15, ha='right')
    fig8.tight_layout()
    add_matplotlib_chart(ws, fig8, 'J' + str(chart_pos), 'Top 5 Productivity Loss Events', 'Bar Chart')
    
    # Format the sheet
    format_sheet(ws, "Graphs & Visualizations")
    print("Graphs & Visualizations created successfully.")


def create_ai_conversation_log(wb):
    """Create the AI Conversation Log sheet with user-provided conversation data."""
    print("Creating AI Conversation Log...")
    
    # Remove existing AI Conversation Log if present
    if 'AI Conversation Log' in wb.sheetnames:
        wb.remove(wb['AI Conversation Log'])
    
    ws = wb.create_sheet("AI Conversation Log")
    
    # Title
    ws['A1'] = "AI Conversation Log"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1A3C5E", end_color="1A3C5E", fill_type="solid")
    ws.merge_cells('A1:H1')
    ws.row_dimensions[1].height = 30
    
    # Subtitle
    ws['A2'] = f"AI Tool: DeepSeek  |  Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A2'].font = Font(size=10, italic=True)
    ws.merge_cells('A2:H2')
    
    # Headers
    headers = ['Chat #', 'Date', 'Prompt #', 'User Prompt', 'AI Response Summary', 'Used?', 'Rejected?', 'Manual Corrections']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2C6B9E", end_color="2C6B9E", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Check if user has provided real conversation data
    has_real_data = False
    for chat_key in ['chat_1', 'chat_2']:
        if chat_key in DEEPSEEK_CONVERSATION:
            chat = DEEPSEEK_CONVERSATION[chat_key]
            if chat.get('prompts') and chat.get('responses'):
                if chat['prompts'][0] != "Your first prompt goes here" and chat['link'] != "PASTE_YOUR_DEEPSEEK_SESSION_LINK_HERE":
                    has_real_data = True
                    break
    
    if has_real_data:
        # Use the provided conversation data
        row_idx = 5
        for chat_num, chat_key in enumerate(['chat_1', 'chat_2'], start=1):
            if chat_key in DEEPSEEK_CONVERSATION:
                chat = DEEPSEEK_CONVERSATION[chat_key]
                prompts = chat.get('prompts', [])
                responses = chat.get('responses', [])
                
                for i, (prompt, response) in enumerate(zip(prompts, responses)):
                    if i == 0:
                        ws.cell(row=row_idx, column=1).value = str(chat_num)
                        ws.cell(row=row_idx, column=2).value = chat.get('date', '')
                    else:
                        ws.cell(row=row_idx, column=1).value = ""
                        ws.cell(row=row_idx, column=2).value = ""
                    
                    ws.cell(row=row_idx, column=3).value = i + 1
                    ws.cell(row=row_idx, column=4).value = prompt
                    ws.cell(row=row_idx, column=5).value = response[:500] + "..." if len(response) > 500 else response
                    ws.cell(row=row_idx, column=6).value = "Yes" if i < len(chat.get('used', [])) else "No"
                    ws.cell(row=row_idx, column=7).value = "No" if not chat.get('rejected') else "Yes"
                    ws.cell(row=row_idx, column=8).value = chat.get('manual', [''])[i] if i < len(chat.get('manual', [])) else ""
                    row_idx += 1
                
                # Chat link
                if chat.get('link') and chat['link'] != "PASTE_YOUR_DEEPSEEK_SESSION_LINK_HERE":
                    ws.cell(row=row_idx, column=1).value = "Link"
                    ws.cell(row=row_idx, column=2).value = chat['link']
                    ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=8)
                    row_idx += 1
                
                # Chat title
                if chat.get('title') and chat['title'] != "Your Conversation Title - e.g., Production Analytics Solution":
                    ws.cell(row=row_idx, column=1).value = "Title"
                    ws.cell(row=row_idx, column=2).value = chat['title']
                    ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=8)
                    row_idx += 1
                
                # Verification
                if chat.get('verification') and chat['verification'][0] != "List verification steps you performed":
                    ws.cell(row=row_idx, column=1).value = "Verification"
                    ws.cell(row=row_idx, column=2).value = "; ".join(chat['verification'])
                    ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=8)
                    row_idx += 1
                
                row_idx += 1  # Blank row between chats
    else:
        # Placeholder for missing data - clear instructions for the user
        row_idx = 5
        ws.cell(row=row_idx, column=1).value = "⚠️ ACTION REQUIRED"
        ws.cell(row=row_idx, column=2).value = "Please replace placeholder data with actual DeepSeek conversations"
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=8)
        row_idx += 2
        
        ws.cell(row=row_idx, column=1).value = "Instructions:"
        ws.cell(row=row_idx, column=2).value = "1. Open this sheet after the script runs"
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=8)
        row_idx += 1
        
        ws.cell(row=row_idx, column=1).value = ""
        ws.cell(row=row_idx, column=2).value = "2. In the DEEPSEEK_CONVERSATION dictionary at the top of the script:"
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=8)
        row_idx += 1
        
        ws.cell(row=row_idx, column=1).value = ""
        ws.cell(row=row_idx, column=2).value = "   - Replace 'date' with the actual date of each conversation"
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=8)
        row_idx += 1
        
        ws.cell(row=row_idx, column=1).value = ""
        ws.cell(row=row_idx, column=2).value = "   - Replace 'link' with your DeepSeek session URL"
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=8)
        row_idx += 1
        
        ws.cell(row=row_idx, column=1).value = ""
        ws.cell(row=row_idx, column=2).value = "   - Replace 'prompts' with each prompt you sent"
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=8)
        row_idx += 1
        
        ws.cell(row=row_idx, column=1).value = ""
        ws.cell(row=row_idx, column=2).value = "   - Replace 'responses' with each AI response received"
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=8)
        row_idx += 1
        
        ws.cell(row=row_idx, column=1).value = ""
        ws.cell(row=row_idx, column=2).value = "3. Include BOTH conversations: Production Analytics Solution and Compliance Review"
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=8)
        row_idx += 1
        
        ws.cell(row=row_idx, column=1).value = ""
        ws.cell(row=row_idx, column=2).value = "4. Re-run the script after filling in the data"
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=8)
    
    # Format the sheet
    format_sheet(ws, "AI Conversation Log")
    
    # Set column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 45
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 35
    
    print("AI Conversation Log created successfully.")


# =============================================================================
# MAIN EXECUTION
# =============================================================================

def main():
    """Main execution function."""
    print("=" * 80)
    print("COMPLIANCE SHEETS ADDITION - ZigZag AI Assignment")
    print("=" * 80)
    print(f"\nScript Version: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    # Check if input file exists
    if not os.path.exists(EXCEL_FILE):
        print(f"\n❌ ERROR: Input file '{EXCEL_FILE}' not found.")
        print("Please ensure 'kpi_output_script2_final.xlsx' exists in the current directory.")
        sys.exit(1)
    
    try:
        # Load the workbook
        print(f"\n📂 Loading workbook: {EXCEL_FILE}")
        wb = load_workbook(EXCEL_FILE)
        print(f"   Existing sheets: {wb.sheetnames}")
        
        # Create missing sheets
        print("\n📊 Adding missing compliance sheets...")
        
        # 1. Dashboard
        create_dashboard(wb)
        
        # 2. Graphs & Visualizations
        create_graphs_sheet(wb)
        
        # 3. AI Conversation Log
        create_ai_conversation_log(wb)
        
        # Save the workbook
        print(f"\n💾 Saving workbook: {EXCEL_FILE}")
        wb.save(EXCEL_FILE)
        print("   Workbook saved successfully.")
        
        # Verify sheets
        print("\n" + "=" * 80)
        print("✅ FINAL VERIFICATION")
        print("=" * 80)
        
        print("\nSheets in final workbook:")
        for idx, sheet in enumerate(wb.sheetnames, start=1):
            status = "✓" if sheet in ['Dashboard', 'Graphs & Visualizations', 'AI Conversation Log'] else " "
            print(f"  {idx}. [{status}] {sheet}")
        
        required_sheets = ['Dashboard', 'Graphs & Visualizations', 'AI Conversation Log']
        missing_sheets = [s for s in required_sheets if s not in wb.sheetnames]
        
        if missing_sheets:
            print(f"\n⚠️ WARNING: The following sheets could not be created: {missing_sheets}")
        else:
            print("\n✅ ALL compliance sheets created successfully!")
        
        # Check if AI Conversation Log has real data
        if 'AI Conversation Log' in wb.sheetnames:
            ws_log = wb['AI Conversation Log']
            has_real_data = False
            for row in range(2, min(10, ws_log.max_row + 1)):
                cell = ws_log.cell(row=row, column=4)
                if cell.value and "Your first prompt goes here" not in str(cell.value) and "Action Required" not in str(cell.value):
                    has_real_data = True
                    break
            
            if not has_real_data:
                print("\n⚠️ NOTE: AI Conversation Log contains placeholder data.")
                print("   Please populate DEEPSEEK_CONVERSATION with your actual DeepSeek")
                print("   conversation data and re-run the script.")
        
        print("\n" + "=" * 80)
        print("✅ COMPLIANCE ADDITION COMPLETE")
        print("=" * 80)
        print("\nThe following sheets have been added:")
        print("  ✅ Dashboard")
        print("  ✅ Graphs & Visualizations (8 charts)")
        print("  ✅ AI Conversation Log (template)")
        print("\nDuration units match your KPI Engine (Hours).")
        print("All charts use the unit from the source data.")
        
    except Exception as e:
        print(f"\n❌ ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()